from collections import Counter
from datetime import timedelta, datetime
from django.conf import settings
from django.db import IntegrityError
from django.http import HttpResponseForbidden
from django.urls import reverse
from django.shortcuts import get_object_or_404, render
from core.decorators import staff_required
from .models import Contact, EmailCampaign, EmailDelivery, EmailClickEvent, EmailTemplate, EmailImage, ContactGroup, EmailCampaignTrackedLink, DEFAULT_FALLBACK_SALUTATION
from django.db.models import OuterRef, Sum, Exists, Prefetch
from django.db.models.deletion import ProtectedError
from media_assets.models import MediaAsset
from .forms import EmailTemplateForm, EmailImageUploadForm, ContactForm, ContactImportForm, ContactGroupForm, SendCampaignForm, NewsletterImageUploadForm
from .scheduling import (find_scheduled_campaign_conflict, get_min_allowed_scheduled_at, get_scheduled_campaign_min_gap_minutes)
from .email_limits import get_daily_email_usage, get_email_count_for_send_form, can_fit_email_count_in_day
from django.contrib import messages
from django.shortcuts import redirect
from openpyxl import load_workbook
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.core.mail import EmailMultiAlternatives, send_mail
from django.template import Context, Template
from django.utils import timezone
import requests
import html
import re
from urllib.parse import quote, unquote, urlparse
from django.db import transaction
from django.core.paginator import Paginator
from django.db.models import Count

# Create your views here.

@staff_required
def dashboard(request):
    cards = [
        {"title": "Šablony", "desc": "Vytvořit a upravit HTML šablony emailů.", "url": reverse("rozesilac:templates")},
        {"title": "Kontakty", "desc": "Správa kontaktů + import z XLSX.", "url": reverse("rozesilac:contacts")},
        {"title": "Odeslat", "desc": "Vybrat šablonu, příjemce a odeslat.", "url": reverse("rozesilac:send")},
        {"title": "Kampaně", "desc": "Historie rozesílek, výsledky a chyby.", "url": reverse("rozesilac:campaigns")},
    ]
    contacts_total = Contact.objects.count()
    active_contacts_total = Contact.objects.filter(is_active=True).count()
    campaigns_total = EmailCampaign.objects.count()
    deliveries_total = EmailDelivery.objects.count()
    sent_deliveries_total = EmailDelivery.objects.filter(status="sent").count()

    campaigns = (EmailCampaign.objects.select_related("template", "created_by").prefetch_related("deliveries__click_events").order_by("-created_at"))

    campaign_stats = []

    for campaign in campaigns:
        deliveries = list(campaign.deliveries.all())

        sent_count = sum(1 for d in deliveries if d.status == "sent")
        failed_count = sum(1 for d in deliveries if d.status == "failed")
        queued_count = sum(1 for d in deliveries if d.status == "queued")

        clicked_delivery_count = 0
        confirmed_unique_click_count_total = 0

        for delivery in deliveries:
            human_events = [e for e in delivery.click_events.all() if not e.is_suspected_bot]
            unique_urls = set(e.original_url for e in human_events)

            if unique_urls:
                clicked_delivery_count += 1
                confirmed_unique_click_count_total += len(unique_urls)

        click_rate_percent = round((clicked_delivery_count / sent_count) * 100, 1) if sent_count else 0
        click_rate_width = max(0, min(100, int(round(click_rate_percent))))

        campaign.sent_count_for_ui = sent_count
        campaign.failed_count_for_ui = failed_count
        campaign.queued_count_for_ui = queued_count
        campaign.clicked_delivery_count_for_ui = clicked_delivery_count
        campaign.confirmed_unique_click_count_total_for_ui = confirmed_unique_click_count_total
        campaign.click_rate_percent_for_ui = click_rate_percent
        campaign.click_rate_width_for_ui = click_rate_width

        campaign_stats.append(campaign)

    latest_campaign = campaign_stats[0] if campaign_stats else None
    top_campaigns = sorted(campaign_stats, key=lambda c: (c.confirmed_unique_click_count_total_for_ui, c.clicked_delivery_count_for_ui, c.created_at,), reverse=True,)[:5]
    recent_campaigns = campaign_stats[:8]

    # --------------------------------------------------
    # Statistiky kontaktů
    # --------------------------------------------------

    contacts = list(Contact.objects.prefetch_related("groups").all())
    contact_emails = [c.email for c in contacts]

    deliveries_for_contacts = (
        EmailDelivery.objects
        .filter(to_email__in=contact_emails)
        .prefetch_related(Prefetch("click_events", queryset=EmailClickEvent.objects.order_by("created_at"),))
    )

    deliveries_by_email = {}
    for d in deliveries_for_contacts:
        deliveries_by_email.setdefault(d.to_email, []).append(d)

    for c in contacts:
        contact_deliveries = deliveries_by_email.get(c.email, [])
        c.delivery_count_for_ui = len(contact_deliveries)

        confirmed_unique_click_count = 0
        clicked_campaigns_count = 0
        last_human_click_at = None

        for d in contact_deliveries:
            human_events = [e for e in d.click_events.all() if not e.is_suspected_bot]
            unique_urls = set(e.original_url for e in human_events)

            confirmed_unique_click_count += len(unique_urls)

            if unique_urls:
                clicked_campaigns_count += 1

            if human_events:
                event_last = human_events[-1].created_at
                if last_human_click_at is None or event_last > last_human_click_at:
                    last_human_click_at = event_last

        c.confirmed_unique_click_count_for_ui = confirmed_unique_click_count
        c.clicked_campaigns_count_for_ui = clicked_campaigns_count
        c.last_human_click_at_for_ui = last_human_click_at

    top_contacts = sorted(
        contacts,
        key=lambda c: (
            c.confirmed_unique_click_count_for_ui,
            c.clicked_campaigns_count_for_ui,
            c.delivery_count_for_ui,
        ),
        reverse=True,
    )[:10]

    max_recent_click_rate = max([c.click_rate_percent_for_ui for c in recent_campaigns], default=0,)

    return render(
        request,
        "rozesilac/dashboard.html",
        {
            "contacts_total": contacts_total,
            "active_contacts_total": active_contacts_total,
            "campaigns_total": campaigns_total,
            "deliveries_total": deliveries_total,
            "sent_deliveries_total": sent_deliveries_total,
            "latest_campaign": latest_campaign,
            "top_campaigns": top_campaigns,
            "top_contacts": top_contacts,
            "recent_campaigns": recent_campaigns,
            "max_recent_click_rate": max_recent_click_rate,
            "cards": cards,
        },
    )


# přechod na media assets pro správu obrázků v rozesílači:

def get_newsletter_image_assets():
    return MediaAsset.objects.filter(
        asset_type=MediaAsset.AssetType.IMAGE,
    ).order_by("-uploaded_at")


def get_newsletter_image_total_size():
    return get_newsletter_image_assets().aggregate(total=Sum("file_size"))["total"] or 0


@staff_required
def templates(request):
    templates = EmailTemplate.objects.order_by("-updated_at", "name")
    return render(request, "rozesilac/templates_list.html", {"templates": templates})

@staff_required
def template_create(request):
    if request.method == "POST":
        form = EmailTemplateForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Šablona byla vytvořena.")
            return redirect("rozesilac:templates")
    else:
        form = EmailTemplateForm()

    recent_images = get_newsletter_image_assets()[:4]
    total_size = get_newsletter_image_total_size()
    limit_size = 500 * 1024 * 1024
    image_upload_form = NewsletterImageUploadForm()

    return render(
        request,
        "rozesilac/template_create.html",
        {
            "form": form,
            "page_title": "Nová šablona",
            "submit_label": "Vytvořit šablonu",
            "recent_images": recent_images,
            "total_size": total_size,
            "limit_size": limit_size,
            "image_upload_form": image_upload_form,
        },
    )

@staff_required
def template_edit(request, template_id):
    template_obj = get_object_or_404(EmailTemplate, id=template_id)

    if request.method == "POST":
        form = EmailTemplateForm(request.POST, instance=template_obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Šablona byla upravena.")
            return redirect("rozesilac:templates")
    else:
        form = EmailTemplateForm(instance=template_obj)

    recent_images = get_newsletter_image_assets()[:4]
    total_size = get_newsletter_image_total_size()
    limit_size = 500 * 1024 * 1024
    image_upload_form = NewsletterImageUploadForm()

    return render(
        request,
        "rozesilac/template_edit.html",
        {
            "form": form,
            "page_title": f"Upravit šablonu: {template_obj.name}",
            "submit_label": "Uložit změny",
            "template_obj": template_obj,
            "recent_images": recent_images,
            "total_size": total_size,
            "limit_size": limit_size,
            "image_upload_form": image_upload_form,
        },
    )

@staff_required
def template_duplicate(request, template_id):
    template_obj = get_object_or_404(EmailTemplate, id=template_id)

    if request.method == "POST":
        base_name = f"{template_obj.name} (kopie)"
        new_name = base_name
        counter = 2

        while EmailTemplate.objects.filter(name=new_name).exists():
            new_name = f"{base_name} {counter}"
            counter += 1

        new_template = EmailTemplate.objects.create(name=new_name, subject=template_obj.subject, preheader=template_obj.preheader, html_body=template_obj.html_body, text_body=template_obj.text_body, fallback_salutation=template_obj.fallback_salutation)

        messages.success(request, f'Šablona byla zduplikována jako "{new_template.name}".')
        return redirect("rozesilac:templates")

    return render(request, "rozesilac/template_duplicate.html", {"template_obj": template_obj},)

@staff_required
def template_delete(request, template_id):
    template_obj = get_object_or_404(EmailTemplate, id=template_id)

    if request.method == "POST":
        try:
            template_obj.delete()
            messages.success(request, "Šablona byla smazána.")
        except ProtectedError:
            messages.error(
                request,
                "Tuto šablonu nelze smazat, protože už byla použita v odeslané kampani. "
                "Potřebujeme, aby historie kampaní zůstala zachována a tohle by nám rozmrdalo ty historické statistiky. Jestli tě existence té šablony sere fakt hodně, napiš mi a nějak to pořešíme :)"
            )
        return redirect("rozesilac:templates")
    return render(request, "rozesilac/template_delete.html", {"template_obj": template_obj},)


def get_contact_salutation(contact, fallback_salutation=DEFAULT_FALLBACK_SALUTATION):
    if contact and contact.salutation and contact.salutation.strip():
        return contact.salutation.strip()

    fallback_salutation = (fallback_salutation or "").strip()
    if fallback_salutation:
        return fallback_salutation

    return DEFAULT_FALLBACK_SALUTATION


@staff_required
def contacts(request):
    if request.method == "POST" and request.POST.get("action") == "delete_contact":
        contact_id = request.POST.get("contact_id")
        Contact.objects.filter(id=contact_id).delete()
        messages.success(request, "Kontakt smazán.")
        return redirect("rozesilac:contacts")

    if request.method == "POST" and request.POST.get("action") == "add_group":
        group_form = ContactGroupForm(request.POST)
        if group_form.is_valid():
            group_form.save()
            messages.success(request, "Skupina byla vytvořena.")
            return redirect("rozesilac:contacts")
    else:
        group_form = ContactGroupForm()

    if request.method == "POST" and request.POST.get("action") == "delete_group":
        group_id = request.POST.get("group_id")
        group = ContactGroup.objects.filter(id=group_id).first()

        if not group:
            messages.error(request, "Skupina nebyla nalezena.")
            return redirect("rozesilac:contacts")

        if group.is_protected:
            messages.error(
                request,
                "Tuto systémovou skupinu nelze smazat. Kontakty v ní ale mazat můžete."
            )
            return redirect("rozesilac:contacts")

        try:
            group.delete()
        except ProtectedError:
            messages.error(
                request,
                "Tuto skupinu nelze smazat, protože je systémově chráněná. Kontakty v ní ale mazat můžete."
            )
            return redirect("rozesilac:contacts")

        messages.success(request, "Skupina byla smazána.")
        return redirect("rozesilac:contacts")

    add_form = ContactForm()
    import_form = ContactImportForm()

    if request.method == "POST" and request.POST.get("action") == "add_contact":
        add_form = ContactForm(request.POST)
        if add_form.is_valid():
            try:
                add_form.save()
                messages.success(request, "Kontakt uložen.")
                return redirect("rozesilac:contacts")
            except IntegrityError:
                add_form.add_error("email", "Tento email už v kontaktech existuje.")

    if request.method == "POST" and request.POST.get("action") == "import":
        import_form = ContactImportForm(request.POST, request.FILES)
        if import_form.is_valid():
            f = import_form.cleaned_data["file"]
            selected_group = import_form.cleaned_data.get("group")

            wb = load_workbook(filename=f, data_only=True)
            ws = wb.active

            header_row = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]

            def find_col(possible_names):
                for i, h in enumerate(header_row):
                    if h in possible_names:
                        return i
                return None

            name_col = find_col({"jméno", "jmeno", "name"})
            email_col = find_col({"email", "e-mail", "e mail", "mail"})
            salutation_col = find_col({"osloveni", "salutation", "pozdrav", "oslovení"})

            if name_col is None or email_col is None:
                messages.error(request, "XLSX musí mít v prvním řádku sloupce 'jméno' a 'email'.")
                return redirect("rozesilac:contacts")

            created = 0
            skipped = 0
            invalid = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                raw_email = row[email_col] if email_col < len(row) else None
                raw_name = row[name_col] if name_col < len(row) else None
                raw_salutation = row[salutation_col] if (salutation_col is not None and salutation_col < len(row)) else None

                email = (str(raw_email).strip() if raw_email is not None else "").lower()
                name = str(raw_name).strip() if raw_name is not None else ""
                salutation = str(raw_salutation).strip() if raw_salutation is not None else ""

                if not email:
                    continue

                try:
                    validate_email(email)
                except ValidationError:
                    invalid += 1
                    continue

                obj, was_created = Contact.objects.get_or_create(email=email, defaults={"name": name, "salutation": salutation, "is_active": True},)

                if was_created:
                    created += 1
                else:
                    skipped += 1

                if selected_group:
                    obj.groups.add(selected_group)

            messages.success(request, f"Import hotový. Přidáno: {created}, přeskočeno (duplicitní): {skipped}, neplatné emaily: {invalid}.")
            return redirect("rozesilac:contacts")

    # --------------------------------------------------
    # Filtrování a stránkování kontaktů
    # --------------------------------------------------

    per_page_options = [20, 50, 100]

    try:
        per_page = int(request.GET.get("per_page", 20))
    except (TypeError, ValueError):
        per_page = 20

    if per_page not in per_page_options:
        per_page = 20

    selected_group_id = request.GET.get("group") or ""

    groups = (
        ContactGroup.objects
        .annotate(contact_count=Count("contacts", distinct=True))
        .order_by("name")
    )

    contacts_qs = (
        Contact.objects
        .prefetch_related("groups")
        .order_by("email")
    )

    selected_group = None

    if selected_group_id:
        selected_group = ContactGroup.objects.filter(id=selected_group_id).first()
        if selected_group:
            contacts_qs = contacts_qs.filter(groups=selected_group)
        else:
            selected_group_id = ""

    paginator = Paginator(contacts_qs, per_page)
    page_obj = paginator.get_page(request.GET.get("page"))

    contacts = list(page_obj.object_list)

    # Query string pro stránkování, aby se při přepínání stránek neztratil filtr
    query_params = request.GET.copy()
    if "page" in query_params:
        del query_params["page"]

    query_string = query_params.urlencode()
    pagination_base_url = f"?{query_string}&" if query_string else "?"

    # --------------------------------------------------
    # Statistiky kontaktů pro přehled v seznamu
    # Důležité: počítáme jen kontakty na aktuální stránce,
    # ne všechny kontakty v databázi.
    # --------------------------------------------------

    contact_emails = [c.email for c in contacts]

    deliveries_for_contacts = (
        EmailDelivery.objects
        .filter(to_email__in=contact_emails)
        .prefetch_related(
            Prefetch(
                "click_events",
                queryset=EmailClickEvent.objects.order_by("created_at"),
            )
        )
    )

    deliveries_by_email = {}

    for d in deliveries_for_contacts:
        deliveries_by_email.setdefault(d.to_email, []).append(d)

    for c in contacts:
        contact_deliveries = deliveries_by_email.get(c.email, [])
        c.delivery_count_for_ui = len(contact_deliveries)

        confirmed_unique_click_count = 0

        for d in contact_deliveries:
            human_events = [
                e for e in d.click_events.all()
                if not e.is_suspected_bot
            ]
            unique_urls = set(e.original_url for e in human_events)
            confirmed_unique_click_count += len(unique_urls)

        c.confirmed_unique_click_count_for_ui = confirmed_unique_click_count

    return render(
        request,
        "rozesilac/contacts.html",
        {
            "contacts": contacts,
            "groups": groups,
            "add_form": add_form,
            "import_form": import_form,
            "group_form": group_form,

            "selected_group_id": str(selected_group_id),
            "selected_group": selected_group,
            "per_page": per_page,
            "per_page_options": per_page_options,
            "page_obj": page_obj,
            "paginator": paginator,
            "pagination_base_url": pagination_base_url,
        },
    )

@staff_required
def contact_detail(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)

    human_clicks_subquery = EmailClickEvent.objects.filter(delivery=OuterRef("pk"), is_suspected_bot=False,)

    bot_clicks_subquery = EmailClickEvent.objects.filter(delivery=OuterRef("pk"), is_suspected_bot=True,)

    deliveries = (
        EmailDelivery.objects
        .filter(to_email=contact.email)
        .select_related("campaign", "campaign__template")
        .order_by("-created_at")
        .annotate(has_human_like_click=Exists(human_clicks_subquery), has_suspected_bot_click=Exists(bot_clicks_subquery),)
        .prefetch_related("campaign__tracked_links", Prefetch("click_events", queryset=EmailClickEvent.objects.order_by("created_at"),))
    )

    total_deliveries = deliveries.count()
    sent_deliveries = deliveries.filter(status="sent").count()
    failed_deliveries = deliveries.filter(status="failed").count()
    clicked_campaigns_count = 0
    confirmed_unique_click_count_total = 0
    last_human_click_at = None

    for delivery in deliveries:
        all_events = list(delivery.click_events.all())
        human_events = [e for e in all_events if not e.is_suspected_bot]

        unique_urls = []
        seen_urls = set()

        for e in human_events:
            if e.original_url not in seen_urls:
                seen_urls.add(e.original_url)
                unique_urls.append(e.original_url)

        delivery.human_click_events_for_ui = human_events
        delivery.clicked_urls_for_ui = unique_urls
        delivery.confirmed_unique_click_count_for_ui = len(unique_urls)
        delivery.first_human_click_at_for_ui = human_events[0].created_at if human_events else None
        delivery.last_human_click_at_for_ui = human_events[-1].created_at if human_events else None

        if unique_urls:
            clicked_campaigns_count += 1
            confirmed_unique_click_count_total += len(unique_urls)

        if human_events:
            candidate_last = human_events[-1].created_at
            if last_human_click_at is None or candidate_last > last_human_click_at:
                last_human_click_at = candidate_last

    return render(
        request,
        "rozesilac/contact_detail.html",
        {
            "contact": contact,
            "deliveries": deliveries,
            "total_deliveries": total_deliveries,
            "sent_deliveries": sent_deliveries,
            "failed_deliveries": failed_deliveries,
            "clicked_campaigns_count": clicked_campaigns_count,
            "confirmed_unique_click_count_total": confirmed_unique_click_count_total,
            "last_human_click_at": last_human_click_at,
        },
    )


@staff_required
def contact_edit(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)

    if request.method == "POST":
        form = ContactForm(request.POST, instance=contact)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Kontakt byl upraven.")
                return redirect("rozesilac:contacts")
            except IntegrityError:
                form.add_error("email", "Tento email už v kontaktech existuje.")
    else:
        form = ContactForm(instance=contact)

    return render(
        request,
        "rozesilac/contact_edit.html",
        {
            "form": form,
            "page_title": f"Upravit kontakt: {contact.email}",
            "submit_label": "Uložit změny",
            "contact_obj": contact,
        },
    )


@staff_required
def images(request):
    upload_form = NewsletterImageUploadForm()

    if request.method == "POST":
        if request.POST.get("action") == "upload":
            upload_form = NewsletterImageUploadForm(request.POST, request.FILES)

            if upload_form.is_valid():
                upload_form.save(uploaded_by=request.user)
                messages.success(request, "Obrázek byl nahrán.")
                return redirect("rozesilac:images")

        elif request.POST.get("action") == "rename":
            image_id = request.POST.get("image_id")
            new_title = (request.POST.get("title") or "").strip()

            image = get_object_or_404(MediaAsset, pk=image_id)
            image.title = new_title
            image.save(update_fields=["title"])

            messages.success(request, "Název obrázku byl upraven.")
            return redirect("rozesilac:images")

        elif request.POST.get("action") == "delete":
            image_id = request.POST.get("image_id")
            obj = get_object_or_404(MediaAsset, id=image_id)

            storage = obj.file.storage if obj.file else None
            file_name = obj.file.name if obj.file else None

            try:
                obj.delete()
            except ProtectedError:
                messages.error(
                    request,
                    "Obrázek nelze smazat, protože se stále používá jinde na webu."
                )
                return redirect("rozesilac:images")

            if storage and file_name:
                transaction.on_commit(lambda: storage.delete(file_name))

            messages.success(request, "Obrázek byl smazán.")
            return redirect("rozesilac:images")

    images = get_newsletter_image_assets()
    total_size = get_newsletter_image_total_size()
    limit_size = 500 * 1024 * 1024

    return render(
        request,
        "rozesilac/images_gallery.html",
        {
            "upload_form": upload_form,
            "images": images,
            "total_size": total_size,
            "limit_size": limit_size,
        },
    )

@staff_required
def image_upload(request):
    if request.method != "POST":
        return HttpResponseForbidden("Pouze POST.")

    form = NewsletterImageUploadForm(request.POST, request.FILES)
    next_url = request.POST.get("next") or "rozesilac:templates"

    if form.is_valid():
        form.save(uploaded_by=request.user)
        messages.success(request, "Obrázek byl nahrán.")
    else:
        for error in form.non_field_errors():
            messages.error(request, error)

        for field_name, errors in form.errors.items():
            if field_name == "__all__":
                continue
            for error in errors:
                messages.error(request, error)

    return redirect(next_url)


# pomocné funkce pro generování absolutních URL, přidání preheaderu do HTML a přepsání odkazů na trackovací redirect.

def add_preheader_to_html(html_content: str, preheader: str) -> str:
    if not html_content or not preheader:
        return html_content

    preheader_html = (
        '<div style="display:none;font-size:1px;color:#fff;line-height:1px;'
        'max-height:0;max-width:0;opacity:0;overflow:hidden;mso-hide:all;">'
        f'{html.escape(preheader)}'
        '&nbsp;&zwnj;&nbsp;&zwnj;&nbsp;&zwnj;&nbsp;&zwnj;&nbsp;&zwnj;&nbsp;'
        '&nbsp;&zwnj;&nbsp;&zwnj;&nbsp;&zwnj;&nbsp;&zwnj;&nbsp;&zwnj;&nbsp;'
        '</div>'
    )

    lowered = html_content.lower()
    body_pos = lowered.find("<body")
    if body_pos != -1:
        body_end = lowered.find(">", body_pos)
        if body_end != -1:
            return html_content[:body_end + 1] + preheader_html + html_content[body_end + 1:]

    return preheader_html + html_content

def is_trackable_url(url: str) -> bool:
    if not url:
        return False

    parsed = urlparse(url.strip())

    if parsed.scheme not in {"http", "https"}:
        return False

    if not parsed.netloc:
        return False

    lowered = url.lower()
    if "unsubscribe" in lowered:
        return False

    return True

def get_client_ip(request):
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded_for:
        return x_forwarded_for.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR")

def is_suspected_bot_click(delivery, user_agent: str, event_time):
    ua = (user_agent or "").lower()

    suspicious_markers = [
        "microsoft",
        "safelinks",
        "defender",
        "exchange",
        "urlscan",
        "crawler",
        "bot",
        "spider",
        "headless",
        "bingpreview",
    ]

    if any(marker in ua for marker in suspicious_markers):
        return True

    if delivery.sent_at:
        diff = (event_time - delivery.sent_at).total_seconds()

        # request ještě před uloženým odesláním = velmi podezřelé
        if diff < 0:
            return True

        # konzervativní okno po odeslání:
        # první technické skeny se často dějí krátce po odeslání,
        # takže radši víc přitvrdíme, než abychom hlásili falešný klik
        if diff <= 30:
            return True

    return False

def find_recent_same_url_click(delivery, target_url, now, window_seconds=30):
    threshold = now - timedelta(seconds=window_seconds)

    return EmailClickEvent.objects.filter(
        delivery=delivery,
        original_url=target_url,
        created_at__gte=threshold,
    ).order_by("-created_at").first()

def has_any_previous_click_for_url(delivery, target_url):
    return EmailClickEvent.objects.filter(
        delivery=delivery,
        original_url=target_url,
    ).exists()

def mark_recent_burst_as_suspicious(delivery, now, window_seconds=8, min_distinct_urls=3):
    """
    Pokud během krátkého okna přišlo pro stejné delivery více různých URL
    ze stejné IP a stejného user-agentu, označíme tyto eventy jako podezřelé.
    """
    threshold = now - timedelta(seconds=window_seconds)

    recent_events = list(
        EmailClickEvent.objects.filter(
            delivery=delivery,
            created_at__gte=threshold,
        ).order_by("created_at")
    )

    groups = {}

    for e in recent_events:
        key = (
            (e.ip_address or "").strip(),
            (e.user_agent or "").strip(),
        )
        groups.setdefault(key, []).append(e)

    changed = False

    for key, events in groups.items():
        distinct_urls = {e.original_url for e in events}
        if len(distinct_urls) >= min_distinct_urls:
            for e in events:
                if not e.is_suspected_bot:
                    e.is_suspected_bot = True
                    e.save(update_fields=["is_suspected_bot"])
                    changed = True

    return changed

def click_tracking(request, token):
    delivery = get_object_or_404(EmailDelivery, tracking_token=token)

    target_url = request.GET.get("url", "").strip()
    target_url = html.unescape(unquote(target_url))

    if not is_trackable_url(target_url):
        return redirect("/")

    now = timezone.now()
    user_agent = request.META.get("HTTP_USER_AGENT", "")
    ip_address = get_client_ip(request)

    had_previous_click_for_url = has_any_previous_click_for_url(delivery, target_url)

    is_duplicate = had_previous_click_for_url

    # Podezřelost určujeme jen z vlastností requestu a času vůči odeslání.
    # NEDĚLÁME závěr podle změny IP/UA proti předchozímu eventu,
    # protože právě ten další event může být reálný uživatel.
    suspected_bot = is_suspected_bot_click(delivery, user_agent, now)

    update_fields = ["click_count"]
    delivery.click_count += 1

    if not delivery.clicked_at:
        delivery.clicked_at = now
        update_fields.append("clicked_at")

    # technickou unikátní URL počítáme jen jednou pro každou cílovou URL
    if not had_previous_click_for_url:
        delivery.unique_click_count += 1
        update_fields.append("unique_click_count")

    delivery.save(update_fields=update_fields)

    created_event = EmailClickEvent.objects.create(
        delivery=delivery,
        original_url=target_url,
        user_agent=user_agent,
        ip_address=ip_address,
        is_suspected_bot=suspected_bot,
        is_duplicate=is_duplicate,
    )

    # --------------------------------------------------
    # Burst detekce:
    # pokud během pár sekund stejné delivery projede více různých URL
    # ze stejné IP a stejného UA, je to téměř jistě scanner.
    # Označíme zpětně celý burst jako podezřelý.
    # --------------------------------------------------
    mark_recent_burst_as_suspicious(
        delivery=delivery,
        now=now,
        window_seconds=8,
        min_distinct_urls=3,
    )

    return redirect(target_url)

def add_click_tracking_to_html(html_content: str, delivery, base_url: str):
    """
    Přepíše všechny <a href="http(s)://..."> odkazy tak,
    aby vedly přes náš tracking redirect.
    Zároveň vrátí seznam původních trackovaných URL.
    """
    if not html_content:
        return html_content, []

    click_base = f"{base_url}{reverse('rozesilac:click_tracking', args=[delivery.tracking_token])}"

    tracked_urls = []

    pattern = re.compile(
        r'(<a\b[^>]*\bhref=)(["\'])(.*?)\2',
        flags=re.IGNORECASE | re.DOTALL,
    )

    def replace_href(match):
        prefix, quote_char, original_href = match.groups()

        original_href = html.unescape(original_href.strip())

        if not is_trackable_url(original_href):
            return match.group(0)

        tracked_urls.append(original_href)

        tracked_url = f"{click_base}?url={quote(original_href, safe='')}"
        return f"{prefix}{quote_char}{tracked_url}{quote_char}"

    rendered_html = pattern.sub(replace_href, html_content)

    # odstraníme duplicity, ale zachováme pořadí
    unique_tracked_urls = list(dict.fromkeys(tracked_urls))

    return rendered_html, unique_tracked_urls

def send_single_delivery(campaign, delivery, base_url: str, from_email: str):
    contact = delivery.contact
    osloveni = get_contact_salutation(contact, fallback_salutation=getattr(campaign, "fallback_salutation", DEFAULT_FALLBACK_SALUTATION))

    if contact:
        unsubscribe_path = reverse("rozesilac:unsubscribe", args=[contact.unsubscribe_token])
        unsubscribe_url = f"{base_url}{unsubscribe_path}"
    else:
        unsubscribe_url = ""

    event_public_url = ""
    vip_event_url = ""

    if campaign.event:
        event_public_path = reverse("events:public_event_detail", args=[campaign.event.slug])
        event_public_url = f"{base_url}{event_public_path}"

        vip_event_path = reverse("events:vip_event_detail", args=[delivery.tracking_token])
        vip_event_url = f"{base_url}{vip_event_path}"

    template_context = Context({
        "osloveni": osloveni,
        "jmeno": contact.name if contact and contact.name else "",
        "email": delivery.to_email,
        "unsubscribe_url": unsubscribe_url,
        "event_public_url": event_public_url,
        "vip_event_url": vip_event_url,
    })

    rendered_subject = Template(campaign.subject).render(template_context)
    rendered_html_body = Template(campaign.html_body).render(template_context)

    preheader = (campaign.preheader or "").strip()
    if preheader:
        rendered_html_body = add_preheader_to_html(rendered_html_body, preheader)

    rendered_html_body, tracked_urls = add_click_tracking_to_html(rendered_html_body, delivery, base_url)

    text_template = campaign.text_body.strip() if campaign.text_body else ""
    if text_template:
        rendered_text_body = Template(text_template).render(template_context)
    else:
        rendered_text_body = "Tento email obsahuje HTML verzi zprávy."

    if preheader:
        rendered_text_body = f"{preheader}\n\n{rendered_text_body}"

    for url in tracked_urls:
        EmailCampaignTrackedLink.objects.get_or_create(campaign=campaign, url=url)

    # DEV = klasický Django email backend
    if settings.APP_ENV != "prod":
        msg = EmailMultiAlternatives(
            subject=rendered_subject,
            body=rendered_text_body,
            from_email=from_email,
            to=[delivery.to_email],
            reply_to=["info@liedersociety.cz"],
        )

        msg.attach_alternative(rendered_html_body, "text/html")
        msg.send(fail_silently=False)

    # PROD = Brevo API
    else:
        to_item = {
            "email": delivery.to_email,
        }

        if delivery.to_name and delivery.to_name.strip():
            to_item["name"] = delivery.to_name.strip()

        payload = {
            "sender": {
                "email": from_email,
                "name": "Lieder Society",
            },
            "to": [to_item],
            "subject": rendered_subject,
            "htmlContent": rendered_html_body,
            "textContent": rendered_text_body,
            "replyTo": {
                "email": "info@liedersociety.cz",
                "name": "Lieder Society",
            },
        }

        headers = {
            "accept": "application/json",
            "api-key": settings.BREVO_API_KEY,
            "content-type": "application/json",
        }

        response = requests.post(
            settings.BREVO_API_URL,
            json=payload,
            headers=headers,
            timeout=20,
        )

        if response.status_code >= 400:
            raise Exception(f"Brevo error {response.status_code}: {response.text}")


def send_campaign_deliveries(campaign, base_url: str, from_email: str):
    sent_count = 0
    failed_count = 0
    limit_reached = False

    campaign.status = "sending"
    campaign.started_at = timezone.now()
    campaign.finished_at = None
    campaign.save(update_fields=["status", "started_at", "finished_at"])

    deliveries = campaign.deliveries.filter(status="queued").order_by("id")

    for delivery in deliveries:
        daily_email_usage = get_daily_email_usage(exclude_campaign_id=campaign.id)

        if daily_email_usage["remaining"] <= 0:
            limit_reached = True
            break

        try:
            send_single_delivery(
                campaign=campaign,
                delivery=delivery,
                base_url=base_url,
                from_email=from_email,
            )

            delivery.status = "sent"
            delivery.sent_at = timezone.now()
            delivery.error = ""
            delivery.save(update_fields=["status", "sent_at", "error"])

            sent_count += 1

        except Exception as exc:
            delivery.status = "failed"
            delivery.error = str(exc)
            delivery.save(update_fields=["status", "error"])

            failed_count += 1

    campaign.finished_at = timezone.now()
    queued_count = campaign.deliveries.filter(status="queued").count()
    total_failed_count = campaign.deliveries.filter(status="failed").count()

    if limit_reached and queued_count > 0:
        tomorrow = timezone.localtime() + timedelta(days=1)
        campaign.status = "scheduled"
        campaign.scheduled_at = tomorrow.replace(hour=8, minute=0, second=0, microsecond=0)

        limit_note = "Odesílání bylo přerušeno kvůli dennímu Brevo limitu. Zbytek zůstává ve frontě."

        current_note = campaign.note or ""

        if limit_note not in current_note:
            campaign.note = ((current_note + " | ") if current_note else "") + limit_note

        campaign.save(update_fields=["status", "finished_at", "scheduled_at", "note"])

    elif queued_count == 0 and total_failed_count == 0:
        campaign.status = "sent"
        campaign.save(update_fields=["status", "finished_at"])

    else:
        campaign.status = "failed"
        campaign.save(update_fields=["status", "finished_at"])

    return sent_count, failed_count

@staff_required
def send(request):
    templates_for_preview = EmailTemplate.objects.all().order_by("name")
    daily_email_usage = get_daily_email_usage()

    running_campaign = (
        EmailCampaign.objects
        .filter(status="sending")
        .order_by("-started_at", "-id")
        .first()
    )

    if request.method == "POST":
        form = SendCampaignForm(request.POST)

        if form.is_valid():
            template = form.cleaned_data["template"]
            send_mode = form.cleaned_data["send_mode"]
            delivery_mode = form.cleaned_data["delivery_mode"]
            scheduled_at = form.cleaned_data.get("scheduled_at")
            test_email = form.cleaned_data.get("test_email")
            contacts = form.cleaned_data.get("contacts")
            note = form.cleaned_data.get("note", "")
            event = form.cleaned_data.get("event")
            from_email = form.cleaned_data.get("from_email") or settings.NEWSLETTER_DEFAULT_FROM_EMAIL

            # --------------------------------------------------
            # Zabránění více okamžitým odesíláním najednou
            # --------------------------------------------------

            if running_campaign:
                started_at_text = (
                    timezone.localtime(running_campaign.started_at).strftime("%d.%m.%Y %H:%M")
                    if running_campaign.started_at
                    else "neznámý čas"
                )

                form.add_error(
                    None,
                    (
                        f"Jiná kampaň (ID {running_campaign.id}) právě odesílá "
                        f"a začala v {started_at_text}. "
                        f"Abyste předešli problémům s Brevo limity, nelze teď vytvořit "
                        f"ani spustit další kampaň. Zkuste to prosím později, až se "
                        f"aktuální kampaň dokončí."
                    ),
                )

                return render(
                    request,
                    "rozesilac/send.html",
                    {
                        "form": form,
                        "templates_for_preview": templates_for_preview,
                        "daily_email_usage": daily_email_usage,
                        "running_campaign": running_campaign,
                    },
                )

            # --------------------------------------------------
            # Brevo daily limit – kontrola před vytvořením kampaně
            # --------------------------------------------------

            daily_email_usage = get_daily_email_usage()
            emails_to_send_count = get_email_count_for_send_form(form.cleaned_data)

            if delivery_mode == "scheduled":
                limit_day = scheduled_at
            else:
                limit_day = timezone.now()

            capacity_check = can_fit_email_count_in_day(
                emails_to_send_count,
                day=limit_day,
            )

            if not capacity_check["ok"]:
                day_label = timezone.localtime(limit_day).strftime("%d.%m.%Y")

                form.add_error(
                    None,
                    (
                        f"Kampaň nelze vytvořit, protože by překročila denní limit Brevo "
                        f"pro den {day_label}. "
                        f"Limit je {capacity_check['limit']} e-mailů. "
                        f"Už odesláno: {capacity_check['sent']}, "
                        f"rezervováno naplánovanými nebo běžícími kampaněmi: {capacity_check['reserved']}, "
                        f"zbývá: {capacity_check['remaining']}, "
                        f"tahle kampaň má {emails_to_send_count} příjemců."
                    ),
                )

                return render(
                    request,
                    "rozesilac/send.html",
                    {
                        "form": form,
                        "templates_for_preview": templates_for_preview,
                        "daily_email_usage": daily_email_usage,
                        "running_campaign": running_campaign,
                    },
                )

            is_test = send_mode == "test"

            if delivery_mode == "scheduled":
                campaign_status = "scheduled"
            else:
                campaign_status = "draft"

            campaign = EmailCampaign.objects.create(
                template=template,
                created_by=request.user,
                subject=template.subject,
                from_email=from_email,
                preheader=template.preheader,
                html_body=template.html_body,
                text_body=template.text_body,
                fallback_salutation=template.fallback_salutation,
                is_test=is_test,
                note=note,
                event=event,
                status=campaign_status,
                scheduled_at=scheduled_at if delivery_mode == "scheduled" else None,
            )

            # --------------------------------------------------
            # Připravíme a uložíme seznam příjemců do deliveries
            # --------------------------------------------------

            if is_test:
                EmailDelivery.objects.create(
                    campaign=campaign,
                    to_email=test_email,
                    to_name="",
                    contact=None,
                    status="queued",
                )
            else:
                for contact in contacts:
                    EmailDelivery.objects.create(
                        campaign=campaign,
                        to_email=contact.email,
                        to_name=contact.name,
                        contact=contact,
                        status="queued",
                    )

            # --------------------------------------------------
            # Scheduled = jen uložit a skončit
            # --------------------------------------------------

            if delivery_mode == "scheduled":
                messages.success(
                    request,
                    f"Kampaň byla naplánována na {timezone.localtime(campaign.scheduled_at):%d.%m.%Y %H:%M}."
                )
                return redirect("rozesilac:campaign_detail", campaign_id=campaign.id)

            # --------------------------------------------------
            # Immediate send
            # --------------------------------------------------

            campaign.status = "scheduled"
            campaign.scheduled_at = timezone.now()
            campaign.save(update_fields=["status", "scheduled_at"])

            messages.success(
                request,
                "Kampaň byla zařazena k okamžitému odeslání. Stav sleduj na detailu kampaně."
            )

            return redirect("rozesilac:campaign_detail", campaign_id=campaign.id)

    else:
        form = SendCampaignForm()

    return render(
        request,
        "rozesilac/send.html",
        {
            "form": form,
            "templates_for_preview": templates_for_preview,
            "daily_email_usage": daily_email_usage,
            "running_campaign": running_campaign,
        },
    )

@staff_required
def campaigns(request):
    campaigns = (EmailCampaign.objects.select_related("created_by", "template", "event").prefetch_related("deliveries").order_by("-created_at"))

    campaign_rows = []

    for campaign in campaigns:
        deliveries = campaign.deliveries.all()
        total_count = deliveries.count()
        sent_count = deliveries.filter(status="sent").count()
        failed_count = deliveries.filter(status="failed").count()
        queued_count = deliveries.filter(status="queued").count()

        campaign_rows.append({
            "campaign": campaign,
            "total_count": total_count,
            "sent_count": sent_count,
            "failed_count": failed_count,
            "queued_count": queued_count,
        })

    return render(request, "rozesilac/campaigns_list.html", {"campaign_rows": campaign_rows},)


@staff_required
def campaign_detail(request, campaign_id):
    campaign = get_object_or_404(EmailCampaign, id=campaign_id)

    human_clicks_subquery = EmailClickEvent.objects.filter(
        delivery=OuterRef("pk"),
        is_suspected_bot=False,
    )

    bot_clicks_subquery = EmailClickEvent.objects.filter(
        delivery=OuterRef("pk"),
        is_suspected_bot=True,
    )

    deliveries = (
        campaign.deliveries.all()
        .order_by("created_at")
        .annotate(
            has_human_like_click=Exists(human_clicks_subquery),
            has_suspected_bot_click=Exists(bot_clicks_subquery),
        )
        .prefetch_related(
            Prefetch(
                "click_events",
                queryset=EmailClickEvent.objects.order_by("created_at"),
            )
        )
    )

    sent_count = deliveries.filter(status="sent").count()
    failed_count = deliveries.filter(status="failed").count()
    queued_count = deliveries.filter(status="queued").count()

    confirmed_clicked_delivery_count = 0
    confirmed_unique_click_count_total = 0
    confirmed_clicked_urls_all = []

    for delivery in deliveries:
        all_events = list(delivery.click_events.all())
        human_events = [e for e in all_events if not e.is_suspected_bot]

        unique_urls = []
        seen_urls = set()

        for e in human_events:
            if e.original_url not in seen_urls:
                seen_urls.add(e.original_url)
                unique_urls.append(e.original_url)

            confirmed_clicked_urls_all.append(e.original_url)

        delivery.human_click_events_for_ui = human_events
        delivery.clicked_urls_for_ui = unique_urls
        delivery.confirmed_unique_click_count_for_ui = len(unique_urls)
        delivery.first_human_click_at_for_ui = human_events[0].created_at if human_events else None
        delivery.last_human_click_at_for_ui = human_events[-1].created_at if human_events else None

        if unique_urls:
            confirmed_clicked_delivery_count += 1
            confirmed_unique_click_count_total += len(unique_urls)

    click_rate_percent = round((confirmed_clicked_delivery_count / sent_count) * 100, 1) if sent_count else 0

    clicked_url_counter = Counter(confirmed_clicked_urls_all)

    tracked_links_stats = []
    for tracked_link in campaign.tracked_links.all():
        tracked_links_stats.append({
            "url": tracked_link.url,
            "click_count": clicked_url_counter.get(tracked_link.url, 0),
        })

    tracked_links_stats.sort(key=lambda x: (-x["click_count"], x["url"]))

    return render(
        request,
        "rozesilac/campaign_detail.html",
        {
            "campaign": campaign,
            "deliveries": deliveries,
            "sent_count": sent_count,
            "failed_count": failed_count,
            "queued_count": queued_count,
            "clicked_delivery_count": confirmed_clicked_delivery_count,
            "total_unique_click_count": confirmed_unique_click_count_total,
            "click_rate_percent": click_rate_percent,
            "tracked_links_stats": tracked_links_stats,
        },
    )

@staff_required
def campaign_cancel(request, campaign_id):
    if request.method != "POST":
        return redirect("rozesilac:campaign_detail", campaign_id=campaign_id)

    campaign = get_object_or_404(EmailCampaign, id=campaign_id)

    if campaign.status != "scheduled":
        messages.error(request, "Zrušit lze jen kampaň, která je právě naplánovaná.")
        return redirect("rozesilac:campaign_detail", campaign_id=campaign.id)

    campaign.status = "cancelled"
    campaign.finished_at = timezone.now()
    campaign.save(update_fields=["status", "finished_at"])

    messages.success(request, "Naplánovaná kampaň byla zrušena.")
    return redirect("rozesilac:campaign_detail", campaign_id=campaign.id)


@staff_required
def campaign_reschedule(request, campaign_id):
    if request.method != "POST":
        return redirect("rozesilac:campaign_detail", campaign_id=campaign_id)

    campaign = get_object_or_404(EmailCampaign, id=campaign_id)

    if campaign.status != "scheduled":
        messages.error(request, "Přeplánovat lze jen kampaň, která je právě naplánovaná.")
        return redirect("rozesilac:campaign_detail", campaign_id=campaign.id)

    raw_scheduled_at = (request.POST.get("scheduled_at") or "").strip()
    if not raw_scheduled_at:
        messages.error(request, "Musíš vyplnit nový datum a čas odeslání.")
        return redirect("rozesilac:campaign_detail", campaign_id=campaign.id)

    try:
        parsed_dt = datetime.strptime(raw_scheduled_at, "%Y-%m-%dT%H:%M")
        scheduled_at = timezone.make_aware(
            parsed_dt,
            timezone.get_current_timezone(),
        )
    except ValueError:
        messages.error(request, "Neplatný formát data a času.")
        return redirect("rozesilac:campaign_detail", campaign_id=campaign.id)

    min_gap_minutes = get_scheduled_campaign_min_gap_minutes()
    min_allowed_scheduled_at = get_min_allowed_scheduled_at()

    if scheduled_at < min_allowed_scheduled_at:
        messages.error(
            request,
            f"Nový čas odeslání musí být alespoň {min_gap_minutes} minut v budoucnosti.",
        )
        return redirect("rozesilac:campaign_detail", campaign_id=campaign.id)

    conflict = find_scheduled_campaign_conflict(
        scheduled_at,
        exclude_campaign_id=campaign.id,
    )

    if conflict:
        conflict_time = timezone.localtime(conflict.scheduled_at).strftime("%d.%m.%Y %H:%M")

        messages.error(
            request,
            (
                f"V okolí tohoto času už je naplánovaná kampaň "
                f"„{conflict.subject}“ na {conflict_time}. "
                f"Mezi naplánovanými kampaněmi musí být aspoň "
                f"{min_gap_minutes} minut rozdíl."
            ),
        )
        return redirect("rozesilac:campaign_detail", campaign_id=campaign.id)

    queued_count = campaign.deliveries.filter(status="queued").count()

    capacity_check = can_fit_email_count_in_day(
        queued_count,
        day=scheduled_at,
        exclude_campaign_id=campaign.id,
    )

    if not capacity_check["ok"]:
        day_label = timezone.localtime(scheduled_at).strftime("%d.%m.%Y")

        messages.error(
            request,
            (
                f"Kampaň nelze přeplánovat na {day_label}, protože by překročila denní limit Brevo. "
                f"Limit je {capacity_check['limit']} e-mailů. "
                f"Už odesláno: {capacity_check['sent']}, "
                f"rezervováno: {capacity_check['reserved']}, "
                f"zbývá: {capacity_check['remaining']}, "
                f"tahle kampaň má ve frontě: {queued_count} příjemců."
            ),
        )
        return redirect("rozesilac:campaign_detail", campaign_id=campaign.id)

    updated_count = EmailCampaign.objects.filter(
        id=campaign.id,
        status="scheduled",
    ).update(
        scheduled_at=scheduled_at,
    )

    if updated_count == 0:
        messages.error(
            request,
            "Kampaň už nelze přeplánovat. Pravděpodobně už není naplánovaná nebo ji právě převzal cron a je zrovna odesílána.",
        )
        return redirect("rozesilac:campaign_detail", campaign_id=campaign.id)

    messages.success(
        request,
        f"Kampaň byla přeplánována na {timezone.localtime(scheduled_at):%d.%m.%Y %H:%M}.",
    )
    return redirect("rozesilac:campaign_detail", campaign_id=campaign.id)


def unsubscribe(request, token):
    contact = get_object_or_404(Contact, unsubscribe_token=token)
    if request.method == "POST":
        if contact.is_active:
            contact.is_active = False
            contact.save()
            send_mail(
                subject="Odhlášení z odběru newsletteru",
                message=f"Tohle je automatická zpráva z Vaňkova super rozesílače. Chci oznámit, že kontakt {contact.email} se odhlásil/a z odběru Lieder newsletteru. V rozesílači bude teď tento kontakt označen jako neaktivní (ale z kontaktů se nesmazal).",
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=["newsletter@liedersociety.cz"],
                fail_silently=True,
            )
        return render(request, "rozesilac/unsubscribe_done.html", {"contact": contact})
    return render(request, "rozesilac/unsubscribe_confirm.html", {"contact": contact})